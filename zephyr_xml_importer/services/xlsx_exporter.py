from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import logging
from typing import Iterable, Iterator, Sequence

try:
    import openpyxl
except Exception as exc:  # pragma: no cover - guarded by runtime dependency
    openpyxl = None
    _OPENPYXL_ERROR = exc
else:  # pragma: no cover - used at runtime
    _OPENPYXL_ERROR = None


logger = logging.getLogger(__name__)

MAX_CELL_LENGTH = 32767

XLSX_HEADERS: list[str] = [
    "Key",
    "Name",
    "Status",
    "Precondition",
    "Objective",
    "Folder",
    "Folder Description",
    "Priority",
    "Labels",
    "Owner",
    "Issues",
    "Test Script (Step-by-step) Step",
    "Test Script (Step-by-step) Test Data",
    "Test Script (Step-by-step) Expected Result",
    "Test Script (Plain Text)",
    "Test Script (BDD)",
]

STEP_NAME_SHEET_TITLE = "TestY Step Names"
STEP_NAME_HEADERS: list[str] = [
    "Key",
    "Case ID",
    "Step Sort Order",
    "Step Name",
]


@dataclass(frozen=True, slots=True)
class ExportStep:
    sort_order: int | None = None
    name: str | None = None
    description: str | None = None
    test_data: str | None = None
    expected_result: str | None = None


@dataclass(frozen=True, slots=True)
class ExportCase:
    case_id: int | None
    key: str | None
    name: str
    status: str | None = None
    precondition: str | None = None
    objective: str | None = None
    folder: str | None = None
    folder_description: str | None = None
    priority: str | None = None
    labels: list[str] = field(default_factory=list)
    owner: str | None = None
    issues: list[str] = field(default_factory=list)
    steps: list[ExportStep] = field(default_factory=list)
    plain_text: str | None = None
    bdd_text: str | None = None


@dataclass(frozen=True, slots=True)
class XlsxExportResult:
    content: bytes
    warnings: list[str] = field(default_factory=list)


def build_xlsx_export(
    cases: Iterable[ExportCase],
    *,
    include_step_names: bool = False,
) -> XlsxExportResult:
    if openpyxl is None:  # pragma: no cover - should be installed via dependencies
        raise RuntimeError("openpyxl is required to export Zephyr XLSX files") from _OPENPYXL_ERROR

    warnings: list[str] = []
    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="Test Cases")
    worksheet.append(list(XLSX_HEADERS))

    step_name_sheet = None
    if include_step_names:
        step_name_sheet = workbook.create_sheet(title=STEP_NAME_SHEET_TITLE)
        step_name_sheet.append(list(STEP_NAME_HEADERS))

    for case in cases:
        _collect_case_warnings(case, warnings)
        for row_values, step_index in _iter_case_rows(case):
            worksheet.append(_truncate_row(row_values, case, warnings, step_index))
        if step_name_sheet is not None and case.steps:
            if not (case.key or "").strip():
                warnings.append(_format_case_warning(case, "Missing key for step-name metadata sheet"))
            for idx, step in enumerate(case.steps):
                sort_order = step.sort_order if step.sort_order is not None else idx
                _append_step_name_row(step_name_sheet, case, sort_order, step, warnings)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return XlsxExportResult(content=buffer.getvalue(), warnings=warnings)


def _collect_case_warnings(case: ExportCase, warnings: list[str]) -> None:
    name = (case.name or "").strip()
    if not name:
        warnings.append("Missing test case name in export payload")
    if case.steps:
        for idx, step in enumerate(case.steps, start=1):
            has_step = bool((step.description or "").strip() or (step.test_data or "").strip())
            has_expected = bool((step.expected_result or "").strip())
            if not has_step and not has_expected:
                warnings.append(_format_case_warning(case, f"Empty step {idx}"))
            elif has_step and not has_expected:
                warnings.append(_format_case_warning(case, f"Empty expected result for step {idx}"))
    elif not ((case.plain_text or "").strip() or (case.bdd_text or "").strip()):
        warnings.append(_format_case_warning(case, "Missing test script content"))


def _format_case_warning(case: ExportCase, message: str) -> str:
    key = (case.key or "").strip()
    if key:
        return f"{message} in case {key}"
    name = (case.name or "").strip()
    if name:
        return f"{message} in case '{name}'"
    return message


def _iter_case_rows(case: ExportCase) -> Iterator[tuple[list[object], int | None]]:
    labels_text = _join_tokens(case.labels)
    issues_text = _join_tokens(case.issues)
    has_steps = bool(case.steps)
    if has_steps:
        for idx, step in enumerate(case.steps):
            base_values = [
                case.key,
                case.name,
                case.status,
                case.precondition,
                case.objective,
                case.folder,
                case.folder_description,
                case.priority,
                labels_text,
                case.owner,
                issues_text,
            ]
            if idx > 0:
                base_values = [None] * len(base_values)
            row_values: list[object] = [
                *base_values,
                step.description,
                step.test_data,
                step.expected_result,
                None,
                None,
            ]
            yield row_values, idx + 1
        return

    row_values = [
        case.key,
        case.name,
        case.status,
        case.precondition,
        case.objective,
        case.folder,
        case.folder_description,
        case.priority,
        labels_text,
        case.owner,
        issues_text,
        None,
        None,
        None,
        case.plain_text,
        case.bdd_text,
    ]
    yield row_values, None


def _join_tokens(items: Sequence[str]) -> str | None:
    cleaned = [item.strip() for item in items if item and item.strip()]
    if not cleaned:
        return None
    return ", ".join(cleaned)


def _truncate_row(
    row: Sequence[object],
    case: ExportCase,
    warnings: list[str],
    step_index: int | None,
) -> list[object]:
    truncated: list[object] = []
    for idx, value in enumerate(row):
        column = XLSX_HEADERS[idx] if idx < len(XLSX_HEADERS) else f"Column {idx + 1}"
        truncated.append(_truncate_cell(value, column, case, warnings, step_index))
    return truncated


def _truncate_cell(
    value: object,
    column: str,
    case: ExportCase,
    warnings: list[str],
    step_index: int | None,
) -> object:
    if value is None:
        return None
    if not isinstance(value, str):
        value = str(value)
    if len(value) <= MAX_CELL_LENGTH:
        return value
    trimmed = value[:MAX_CELL_LENGTH]
    warning = _format_case_warning(case, f"Value too long for '{column}'")
    if step_index is not None:
        warning = f"{warning} (step {step_index})"
    warnings.append(warning)
    return trimmed


def _append_step_name_row(
    sheet: object,
    case: ExportCase,
    sort_order: int,
    step: ExportStep,
    warnings: list[str],
) -> None:
    key = (case.key or "").strip() or None
    case_id = case.case_id
    name = (step.name or "").strip()

    if not name:
        warnings.append(_format_case_warning(case, f"Empty step name for sort_order {sort_order}"))
        name_value: str | None = None
    else:
        max_length = 255
        if len(name) > max_length:
            warnings.append(
                _format_case_warning(
                    case,
                    f"Step name too long ({len(name)} > {max_length}); truncating",
                )
            )
            name = name[:max_length]
        name_value = name

    row = [key, case_id, sort_order, name_value]
    try:
        sheet.append(row)
    except Exception:  # pragma: no cover - defensive for openpyxl differences
        return
