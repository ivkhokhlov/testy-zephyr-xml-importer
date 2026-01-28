from __future__ import annotations

from contextlib import contextmanager
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
import re
from typing import Any, BinaryIO, Iterator

try:
    import openpyxl
except Exception as exc:  # pragma: no cover - guarded by runtime dependency
    openpyxl = None
    _OPENPYXL_ERROR = exc
else:  # pragma: no cover - used at runtime
    _OPENPYXL_ERROR = None

from .models import ZephyrFolder, ZephyrIssue, ZephyrStep, ZephyrTestCase


HEADER_KEY = "key"
HEADER_NAME = "name"
HEADER_STATUS = "status"
HEADER_PRECONDITION = "precondition"
HEADER_OBJECTIVE = "objective"
HEADER_FOLDER = "folder"
HEADER_FOLDER_DESCRIPTION = "folder_description"
HEADER_PRIORITY = "priority"
HEADER_LABELS = "labels"
HEADER_OWNER = "owner"
HEADER_ISSUES = "issues"
HEADER_STEP = "step"
HEADER_STEP_TEST_DATA = "step_test_data"
HEADER_STEP_EXPECTED = "step_expected"
HEADER_PLAIN_TEXT = "plain_text"
HEADER_BDD = "bdd"


@contextmanager
def _open_workbook(source: str | Path | BinaryIO | bytes) -> Iterator[Any]:
    if openpyxl is None:  # pragma: no cover - should be installed via dependencies
        raise RuntimeError("openpyxl is required to parse Zephyr XLSX exports") from _OPENPYXL_ERROR

    if isinstance(source, (bytes, bytearray)):
        buffer = BytesIO(bytes(source))
        wb = openpyxl.load_workbook(buffer, read_only=False, data_only=True)
    elif isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(source, read_only=False, data_only=True)
    else:
        stream = source
        seekable = False
        try:
            seekable = bool(getattr(stream, "seekable", lambda: False)())
        except Exception:
            seekable = False
        if seekable:
            try:
                stream.seek(0)
            except Exception:
                seekable = False
        if seekable:
            wb = openpyxl.load_workbook(stream, read_only=False, data_only=True)
        else:
            data = stream.read()
            if isinstance(data, str):
                data = data.encode("utf-8")
            buffer = BytesIO(data)
            wb = openpyxl.load_workbook(buffer, read_only=False, data_only=True)
    try:
        yield wb
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _coerce_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    cleaned = str(value).strip()
    return cleaned if cleaned else None


def _split_tokens(value: Any) -> list[str]:
    text = _coerce_text(value)
    if not text:
        return []
    normalized = re.sub(r"[\n\r;]+", ",", text)
    tokens = [t.strip() for t in normalized.split(",") if t.strip()]
    seen: set[str] = set()
    cleaned: list[str] = []
    for token in tokens:
        compact = " ".join(token.split())
        if not compact or compact in seen:
            continue
        seen.add(compact)
        cleaned.append(compact)
    return cleaned


def _header_role(normalized: str) -> str | None:
    if normalized == "key":
        return HEADER_KEY
    if normalized == "name":
        return HEADER_NAME
    if normalized == "status":
        return HEADER_STATUS
    if normalized == "precondition":
        return HEADER_PRECONDITION
    if normalized == "objective":
        return HEADER_OBJECTIVE
    if normalized == "folder":
        return HEADER_FOLDER
    if normalized == "folderdescription":
        return HEADER_FOLDER_DESCRIPTION
    if normalized == "priority":
        return HEADER_PRIORITY
    if normalized in {"labels", "label"}:
        return HEADER_LABELS
    if normalized == "owner":
        return HEADER_OWNER
    if normalized == "issues" or "coverageissues" in normalized or "coverageissue" in normalized:
        return HEADER_ISSUES
    if "testscript" in normalized and "stepbystep" in normalized and normalized.endswith("step"):
        return HEADER_STEP
    if "testscript" in normalized and "stepbystep" in normalized and "testdata" in normalized:
        return HEADER_STEP_TEST_DATA
    if "testscript" in normalized and "stepbystep" in normalized and "expected" in normalized:
        return HEADER_STEP_EXPECTED
    if "testscript" in normalized and ("plaintext" in normalized or "plain" in normalized):
        return HEADER_PLAIN_TEXT
    if "testscript" in normalized and "bdd" in normalized:
        return HEADER_BDD
    return None


def _build_header_index(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_index: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        normalized = _normalize_header(raw)
        if not normalized:
            continue
        role = _header_role(normalized)
        if role and role not in header_index:
            header_index[role] = idx
    return header_index


def _row_has_any_value(row: tuple[Any, ...]) -> bool:
    for value in row:
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip():
                return True
            continue
        return True
    return False


def _row_value(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None:
        return None
    if index < 0 or index >= len(row):
        return None
    return row[index]


@dataclass
class _XlsxCaseBuilder:
    key: str | None
    name: str | None
    status: str | None
    precondition: str | None
    objective: str | None
    folder: str | None
    folder_description: str | None
    priority: str | None
    owner: str | None
    labels: list[str] = field(default_factory=list)
    issues: list[ZephyrIssue] = field(default_factory=list)
    test_script_text: str | None = None
    test_script_type: str | None = None
    steps: list[ZephyrStep] = field(default_factory=list)

    def add_step_from_row(self, row: tuple[Any, ...], header_index: dict[str, int]) -> None:
        step_value = _coerce_text(_row_value(row, header_index.get(HEADER_STEP)))
        test_data = _coerce_text(_row_value(row, header_index.get(HEADER_STEP_TEST_DATA)))
        expected = _coerce_text(_row_value(row, header_index.get(HEADER_STEP_EXPECTED)))
        if not any([step_value, test_data, expected]):
            return
        index = len(self.steps)
        self.steps.append(
            ZephyrStep(
                index=index,
                description=step_value,
                test_data=test_data,
                expected_result=expected,
            )
        )

    def update_from_row(self, row: tuple[Any, ...], header_index: dict[str, int]) -> None:
        if not self.labels:
            self.labels = _split_tokens(_row_value(row, header_index.get(HEADER_LABELS)))
        if not self.issues:
            self.issues = _build_issues(_row_value(row, header_index.get(HEADER_ISSUES)))
        if not self.folder:
            self.folder = _coerce_text(_row_value(row, header_index.get(HEADER_FOLDER)))
        if not self.folder_description:
            self.folder_description = _coerce_text(
                _row_value(row, header_index.get(HEADER_FOLDER_DESCRIPTION))
            )
        if not self.test_script_text:
            self._set_script_text(row, header_index)

    def _set_script_text(self, row: tuple[Any, ...], header_index: dict[str, int]) -> None:
        plain_text = _coerce_text(_row_value(row, header_index.get(HEADER_PLAIN_TEXT)))
        bdd_text = _coerce_text(_row_value(row, header_index.get(HEADER_BDD)))
        if plain_text:
            self.test_script_text = plain_text
            self.test_script_type = "plain"
        elif bdd_text:
            self.test_script_text = bdd_text
            self.test_script_type = "bdd"

    def to_test_case(self) -> ZephyrTestCase:
        test_script_type = self.test_script_type
        test_script_text = self.test_script_text
        if self.steps:
            test_script_type = "steps"
            test_script_text = None
        return ZephyrTestCase(
            zephyr_id=None,
            key=self.key,
            name=self.name,
            folder=self.folder,
            folder_description=self.folder_description,
            objective=self.objective,
            precondition=self.precondition,
            status=self.status,
            priority=self.priority,
            owner=self.owner,
            created_by=None,
            created_on=None,
            updated_by=None,
            updated_on=None,
            param_type=None,
            parameters=[],
            test_data_wrapper=None,
            labels=self.labels,
            issues=self.issues,
            attachments=[],
            test_script_type=test_script_type,
            test_script_text=test_script_text,
            steps=self.steps,
        )


def _build_issues(value: Any) -> list[ZephyrIssue]:
    return [ZephyrIssue(key=issue) for issue in _split_tokens(value)]


def iter_test_cases_xlsx(source: str | Path | BinaryIO | bytes) -> Iterator[ZephyrTestCase]:
    with _open_workbook(source) as workbook:
        ws = workbook.active
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return

        header_index = _build_header_index(tuple(header_row))
        current: _XlsxCaseBuilder | None = None

        for row in rows:
            row_tuple = tuple(row)
            if not _row_has_any_value(row_tuple):
                continue

            key_value = _coerce_text(_row_value(row_tuple, header_index.get(HEADER_KEY)))
            if key_value:
                if current is not None:
                    yield current.to_test_case()

                current = _XlsxCaseBuilder(
                    key=key_value,
                    name=_coerce_text(_row_value(row_tuple, header_index.get(HEADER_NAME))),
                    status=_coerce_text(_row_value(row_tuple, header_index.get(HEADER_STATUS))),
                    precondition=_coerce_text(
                        _row_value(row_tuple, header_index.get(HEADER_PRECONDITION))
                    ),
                    objective=_coerce_text(
                        _row_value(row_tuple, header_index.get(HEADER_OBJECTIVE))
                    ),
                    folder=_coerce_text(_row_value(row_tuple, header_index.get(HEADER_FOLDER))),
                    folder_description=_coerce_text(
                        _row_value(row_tuple, header_index.get(HEADER_FOLDER_DESCRIPTION))
                    ),
                    priority=_coerce_text(_row_value(row_tuple, header_index.get(HEADER_PRIORITY))),
                    owner=_coerce_text(_row_value(row_tuple, header_index.get(HEADER_OWNER))),
                    labels=_split_tokens(_row_value(row_tuple, header_index.get(HEADER_LABELS))),
                    issues=_build_issues(_row_value(row_tuple, header_index.get(HEADER_ISSUES))),
                )
                current._set_script_text(row_tuple, header_index)
                current.add_step_from_row(row_tuple, header_index)
            else:
                if current is None:
                    continue
                current.update_from_row(row_tuple, header_index)
                current.add_step_from_row(row_tuple, header_index)

        if current is not None:
            yield current.to_test_case()


def build_folders_from_cases(cases: list[ZephyrTestCase]) -> dict[str, ZephyrFolder]:
    folders: dict[str, ZephyrFolder] = {}
    for tc in cases:
        folder_path = (tc.folder or "").strip()
        if not folder_path:
            continue
        folder_description = _coerce_text(tc.folder_description)
        existing = folders.get(folder_path)
        if existing is None:
            folders[folder_path] = ZephyrFolder(
                full_path=folder_path,
                index=None,
                description=folder_description,
            )
        elif existing.description is None and folder_description:
            folders[folder_path] = ZephyrFolder(
                full_path=existing.full_path,
                index=existing.index,
                description=folder_description,
            )
    return folders
