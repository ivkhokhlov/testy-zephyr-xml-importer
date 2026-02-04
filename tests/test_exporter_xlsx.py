from __future__ import annotations

from pathlib import Path

import pytest

try:
    import openpyxl
except Exception:  # pragma: no cover - dependency should be installed in runtime
    openpyxl = None

from zephyr_xml_importer.services.xlsx_exporter import (
    ExportCase,
    ExportStep,
    build_xlsx_export,
)
from zephyr_xml_importer.services.xlsx_parser import iter_test_cases_xlsx


@pytest.mark.skipif(openpyxl is None, reason="openpyxl is required for XLSX export")
def test_export_xlsx_roundtrip(tmp_path: Path) -> None:
    cases = [
        ExportCase(
            case_id=1,
            key="ES-T1",
            name="Login works",
            status="Approved",
            precondition="User exists",
            objective="Goal",
            folder="/ui/Login",
            folder_description="Login form checks",
            priority="High",
            labels=["smoke", "regression"],
            owner="alice",
            issues=["ES-1", "ES-2"],
            steps=[
                ExportStep(
                    title="Open login",
                    description="Open page",
                    test_data="user=admin",
                    expected_result="Page opens",
                ),
                ExportStep(
                    title="Submit login",
                    description="Submit form",
                    test_data=None,
                    expected_result="Dashboard opens",
                ),
            ],
            plain_text=None,
            bdd_text=None,
        ),
        ExportCase(
            case_id=2,
            key=None,
            name="Plain case",
            status="Draft",
            precondition=None,
            objective=None,
            folder="",
            folder_description=None,
            priority="Low",
            labels=["label1"],
            owner="bob",
            issues=[],
            steps=[],
            plain_text="Scenario text",
            bdd_text=None,
        ),
    ]

    result = build_xlsx_export(cases, include_extra_testy_fields=True)
    workbook_path = tmp_path / "export.xlsx"
    workbook_path.write_bytes(result.content)

    parsed = list(iter_test_cases_xlsx(workbook_path))
    assert len(parsed) == 2

    first = parsed[0]
    assert first.key == "ES-T1"
    assert first.name == "Login works"
    assert first.folder == "/ui/Login"
    assert first.folder_description == "Login form checks"
    assert first.labels == ["smoke", "regression"]
    assert [issue.key for issue in first.issues] == ["ES-1", "ES-2"]
    assert first.test_script_type == "steps"
    assert len(first.steps) == 2
    assert first.steps[0].title == "Open login"
    assert first.steps[0].description == "Open page"
    assert first.steps[0].test_data == "user=admin"
    assert first.steps[0].expected_result == "Page opens"
    assert first.steps[1].title == "Submit login"
    assert first.steps[1].description == "Submit form"
    assert first.steps[1].expected_result == "Dashboard opens"

    second = parsed[1]
    assert second.key is None
    assert second.name == "Plain case"
    assert second.test_script_type == "plain"
    assert second.test_script_text == "Scenario text"


@pytest.mark.skipif(openpyxl is None, reason="openpyxl is required for XLSX export")
def test_export_xlsx_header_extra_fields_toggle(tmp_path: Path) -> None:
    cases = [
        ExportCase(
            case_id=1,
            key="ES-T1",
            name="Header check",
        )
    ]

    base_result = build_xlsx_export(cases)
    base_path = tmp_path / "base.xlsx"
    base_path.write_bytes(base_result.content)
    base_wb = openpyxl.load_workbook(base_path, read_only=True, data_only=True)
    base_headers = [cell for cell in next(base_wb.active.iter_rows(values_only=True))]
    base_wb.close()
    assert "Step Name" not in base_headers

    extra_result = build_xlsx_export(cases, include_extra_testy_fields=True)
    extra_path = tmp_path / "extra.xlsx"
    extra_path.write_bytes(extra_result.content)
    extra_wb = openpyxl.load_workbook(extra_path, read_only=True, data_only=True)
    extra_headers = [cell for cell in next(extra_wb.active.iter_rows(values_only=True))]
    extra_wb.close()
    assert "Step Name" in extra_headers
